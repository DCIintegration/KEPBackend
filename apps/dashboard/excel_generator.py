from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import os

class ExcelGenerator:
    def __init__(self, log_function):
        self.log = log_function
    
    def generate_excel(self, output_file, horas_por_ot, horas_mensuales_por_ot, info_por_ot):
        """
        Genera el archivo Excel con una sola tabla que incluye horas totales y mensuales.
        Solo se incluyen las OTs que tienen horas mensuales reportadas.
        """
        self.log("Generando archivo Excel...")

        # Crear un archivo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Horas por OT"

        # Encabezados de la tabla
        encabezados = [
            "OT", "Cliente", 
            "Horas totales", "Horas en planta (totales)", "Horas en oficina (totales)", 
            "Horas mensuales totales", "Horas mensuales en planta", "Horas mensuales en oficina"
        ]
        ws.append(encabezados)

        # Aplicar formato a la fila de encabezados
        for cell in ws[1]:  # Los encabezados están en la primera fila
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Procesar las OTs que tienen horas mensuales
        for ot, data_mensual in horas_mensuales_por_ot.items():
            if ot in horas_por_ot:  # Verificar si la OT tiene horas totales
                data_total = horas_por_ot[ot]
                cliente = info_por_ot[ot]["cliente"]

                # Calcular horas totales
                horas_totales = data_total["total"]
                horas_planta_total = data_total.get("planta", 0)
                horas_oficina_total = horas_totales - horas_planta_total

                # Calcular horas mensuales
                horas_mensuales = data_mensual["total"]
                horas_planta_mensual = data_mensual.get("planta", 0)
                horas_oficina_mensual = horas_mensuales - horas_planta_mensual

                # Añadir fila al Excel
                ws.append([
                    ot, cliente,
                    horas_totales, horas_planta_total, horas_oficina_total,
                    horas_mensuales, horas_planta_mensual, horas_oficina_mensual
                ])

        # Ajustar el ancho de las columnas
        for column in range(1, ws.max_column + 1):
            column_letter = get_column_letter(column)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Guardar el archivo de Excel
        try:
            wb.save(output_file)
            self.log(f"Archivo de Excel '{output_file}' generado correctamente.")
        except Exception as e:
            self.log(f"Error al guardar el archivo Excel: {str(e)}")
            # Intentar guardar en una ubicación alternativa
            alternative_path = os.path.join(os.path.dirname(__file__), "Horas_por_OT.xlsx")
            wb.save(alternative_path)
            self.log(f"Archivo guardado en ubicación alternativa: {alternative_path}")